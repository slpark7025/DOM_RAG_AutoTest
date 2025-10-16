# -*- coding: utf-8 -*-
# pip install -U langchain langchain-openai langchain-community chromadb tiktoken python-dotenv openpyxl

# =============================================================================
# 본 파일은 "시나리오 → URL 스코프 기반 RAG → (유사 시나리오 코드 재활용) → LLM 코드 생성
# → 후처리·검증·보정 → 실행·로깅 → 시나리오-스크립트 적합성 평가"를 일괄 수행하는 파이프라인입니다.
#
# 주석 가이드:
# - 각 함수 앞에는 "무엇을/왜/어떻게" 실행하는지 상세 설명을 기술했습니다.
# - 복잡한 로직(정규식, AST 교체, RAG 검색, Step 매칭 등) 내부에는 단계별 주석을 추가했습니다.
# - 기존 로직/코드는 변경하지 않았고, 설명 주석만 추가했습니다.
# =============================================================================

# ========== [IMPORTS] ==========
# Standard Library
import json
import math
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from difflib import SequenceMatcher
from urllib.parse import urljoin, urlsplit

# Third-Party
import chromadb
from dotenv import load_dotenv
from langchain_community.vectorstores import Chroma
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnableMap
from langchain_openai import ChatOpenAI, OpenAIEmbeddings

# Local
from validate_selector_ids import validate_generated_code  # ID 정합성 검증 모듈


# ========== [CONFIG] ==========
CHROMA_BASE_DIR = "./chroma"  # JSON별 DB 폴더들이 위치한 루트
K_PER_BASE = 200  # 200~400 사이로 조절해서 사용 - 현재 200, 300개 각각 테스트함


# ========================= [UTILS: PATH/JSON/TEXT] =========================
def derive_base_path(s: str) -> str:
    """전체 URL/상대경로 무엇이든 받아 /vpes/<section> 형태로 변환
    - 목적: 다양한 입력(URL 전체/상대경로 등)에서 테스트 스코프의 기준이 되는 base path를 추출.
    - 방법:
      1) urlsplit으로 path만 추출
      2) path를 '/'로 나누고 빈 세그먼트 제거
      3) 앞 2개 세그먼트를 조합해 '/vpes/<section>' 같은 형태로 맞춤(>=2개일 때)
    - 예: 'http://host/a/b/c' → '/a/b'
    """
    def to_base_path(path: str) -> str:
        parts = [seg for seg in path.split("/") if seg]
        return ("/" + "/".join(parts[:2])) if len(parts) >= 2 else path
    return to_base_path(urlsplit(s).path)


def extract_keywords(text):
    """문장 내에서 검색/필터링에 유용한 키워드 후보를 추출
    - 역할: 한글/영문/숫자/언더스코어/대시 문자열 + 대괄호[...] + 괄호(...)의 내용을 분리 추출
    - 사용처: 컨텍스트 소스파일 선별 및 대괄호 힌트 강화 등
    """
    words = [w.lower() for w in re.findall(r'[A-Za-z0-9가-힣_-]+', text)]
    bracket_hints = [h.lower() for h in re.findall(r'\[([^\]]+)\]', text)]
    paren_hints   = [h.lower() for h in re.findall(r'\(([^)]+)\)', text)]
    return words + bracket_hints + paren_hints


def _extract_first_json(s: str) -> str | None:
    """임의의 문자열에서 첫 번째 '유효한 JSON'을 찾아 반환
    - 1) 코드펜스 내부 블록이 JSON이면 그대로 사용
    - 2) 전체 문자열을 스캔하여 가장 바깥쪽 괄호 깊이를 추적하며 첫 유효 JSON 객체를 검증 반환
    - 실패 시 None
    """
    s = s.strip()

    # 1) 코드펜스(언어태그 유무 무관) 우선 탐색
    m = re.search(r"```(?:\s*[a-zA-Z0-9_-]+)?\s*\n?([\s\S]*?)\n?```", s)
    if m:
        block = m.group(1).strip()
        if block.startswith("{") or block.startswith("["):
            try:
                json.loads(block)
                return block
            except Exception:
                pass

    # 2) 괄호 깊이로 { ... } 블록을 추적하면서 유효성 검사
    depth = 0; start = -1
    for i, ch in enumerate(s):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            if depth > 0:
                depth -= 1
                if depth == 0 and start != -1:
                    cand = s[start:i+1]
                    try:
                        json.loads(cand)
                        return cand
                    except Exception:
                        start = -1
    return None


def remove_markdown(generated_code: str) -> str:
    """코드펜스 마크다운 제거
    - LLM 출력이 ```python ... ``` 형태일 때 코드만 추출해 후처리하기 위한 정리 유틸
    """
    return generated_code.replace("```python", "").replace("```", "").strip()


def ensure_import(g: str, line: str) -> str:
    """특정 import 라인이 없으면 상단에 삽입
    - 단순 포함 체크(line in g) 방식으로, 동일한 기능의 다른 import 변형은 감지하지 않음(의도된 간결성)
    """
    if line in g:
        return g
    lines = g.splitlines(True)
    insert_at = 0
    if lines and (lines[0].startswith("#!") or "coding" in lines[0]):
        insert_at = 1
    return "".join(lines[:insert_at] + [line + "\n"] + lines[insert_at:])


# ========================= [UTILS: CODE PATCH/POST-PROC] =========================
def insert_sleep_before_assert(generated_code: str) -> str:
    """첫 assert 앞에 sleep(2)를 1회 삽입(안티-플레이키)
    - 목적: 렌더 지연 등으로 인한 비결정적 실패 감소
    - 로직: 멀티라인 정규식으로 첫 assert/self.assert/pytest assert 위치를 찾아 그 직전에 sleep 추가
    - 'from time import sleep' 누락 시 자동 삽입
    """
    pat = re.compile(r'^(\s*)(?:self\.)?assert', re.MULTILINE)
    m = pat.search(generated_code)
    if not m:
        return generated_code
    lead = m.group(1)
    g = generated_code[:m.start()] + f"{lead}sleep(2)\n" + generated_code[m.start():]
    if "sleep(" in g and "from time import sleep" not in g:
        g = "from time import sleep\n" + g
    return g


def patch_unittest_main(generated_code: str) -> str:
    """unittest 진입부(main) 인자/exit 동작 표준화
    - 목적: 노트북/런타임과의 호환(첫 인자 무시, exit=False) 확보
    - 방법: '__main__' 블록의 unittest.main() 호출 패턴 치환
    """
    return re.sub(
        r'if __name__ == ["\']__main__["\']:\s*unittest\.main\(\)',
        'if __name__ == "__main__":\n    unittest.main(argv=[\'first-arg-is-ignored\'], exit=False)',
        generated_code
    )


def patch_teardown(generated_code: str) -> str:
    """
    tearDown(self) 내용을 '표준 블록'으로 완전히 교체하되,
    함수 경계를 정확히 산정해 중복 self.driver.quit()이나 잔여 주석/try 블록이 남지 않도록 한다.
    - __main__ 블록 등 함수 밖 코드는 절대 건드리지 않음
    - tearDown이 없으면 unittest.TestCase 클래스 직후에 삽입
    - 각 행 사이 빈 줄(공백 행) 없이 출력

    구현 상세:
    1) 우선 AST 파싱으로 def tearDown 위치/범위를 정확히 탐지한 뒤 교체
    2) AST 실패 시 정규식 fallback: def 라인과 그보다 깊은 들여쓰기 라인들을 '한 블록'으로 간주해 교체
    3) tearDown 자체가 없는 경우, unittest.TestCase 상속 클래스 선언 직후에 표준 블록 삽입
    """
    import ast

    # 표준 본문(빈 줄 없음)
    def _make_teardown(indent: str) -> str:
        body = (
            f"{indent}def tearDown(self):\n"
            f"{indent}    result = default_setting.get_result(self)\n"
            f"{indent}    default_setting.upload_result(self, case_id, result)\n"
            f"{indent}    self.driver.quit()\n"
        )
        return body

    code = generated_code
    uses_crlf = "\r\n" in code
    # 토크나이즈 안정화를 위해 탭->4공백만 변환(출력은 원래 개행 유지)
    code_norm = code.replace("\t", "    ")

    # 1) AST로 정확히 tearDown 함수 범위를 잡아 교체
    try:
        tree = ast.parse(code_norm)
        lines = code_norm.splitlines(keepends=True)

        # tearDown 함수 노드 찾기
        td = None
        for node in ast.walk(tree):
            if isinstance(node, ast.FunctionDef) and node.name == "tearDown":
                td = node
                break

        if td and hasattr(td, "lineno") and hasattr(td, "end_lineno") and td.end_lineno is not None:
            start = td.lineno - 1
            end = td.end_lineno  # slice용으로 이미 +1된 느낌으로 사용
            # def 라인의 실제 들여쓰기 추출
            def_line = lines[start]
            base_indent = def_line[: len(def_line) - len(def_line.lstrip(" "))]

            # 교체
            new_block = _make_teardown(base_indent)
            new_lines = lines[:start] + [new_block] + lines[end:]
            out = "".join(new_lines)
            if uses_crlf:
                out = out.replace("\n", "\r\n")
            return out
    except SyntaxError:
        # AST 실패 시 아래 정규식 fallback으로 처리
        pass

    # 2) 정규식 fallback: def tearDown 라인과 그 '더 깊은 들여쓰기 라인들' 전체를 교체
    pat = re.compile(
        r'^([ \t]*)def\s+tearDown\s*\(\s*self\s*\)\s*:\s*\r?\n'   # def 라인
        r'('
        r'(?:\1[ \t]+.*\r?\n|'                                    # 더 깊은 들여쓰기 라인
        r'[ \t]*\r?\n)*'                                          # 혹은 빈 줄들
        r')',
        flags=re.MULTILINE,
    )

    def _repl(m: re.Match) -> str:
        base = m.group(1).replace("\t", "    ")
        return _make_teardown(base)

    code2, n = pat.subn(_repl, code_norm, count=1)
    if n > 0:
        if uses_crlf:
            code2 = code2.replace("\n", "\r\n")
        return code2

    # 3) tearDown이 없으면: unittest.TestCase 클래스 선언 직후에 삽입
    class_pat = re.compile(
        r'^([ \t]*)class\s+[A-Za-z_][A-Za-z0-9_]*\s*\(\s*unittest\.TestCase\s*\)\s*:\s*\r?\n',
        flags=re.MULTILINE,
    )
    m = class_pat.search(code_norm)
    if m:
        cls_indent = m.group(1).replace("\t", "    ")
        body_indent = cls_indent + "    "
        insert_at = m.end()
        out = code_norm[:insert_at] + _make_teardown(body_indent) + code_norm[insert_at:]
        if uses_crlf:
            out = out.replace("\n", "\r\n")
        return out

    # 4) 클래스도 없으면 원본 유지
    return generated_code


def strip_document_prefix_xpaths(g: str) -> str:
    """
    저장된 DOM XPath에 '/[document]' 프리픽스가 있을 때 Selenium 호환을 위해 제거
    - 대상: (By.XPATH, "..."), .find_element(By.XPATH,"..."), .find_elements(By.XPATH,"...")
    - 내부 치환은 따옴표 안의 XPath 문자열 앞쪽 '/[document][n]/' 패턴만 제거
    - 예) '/[document][1]/html[1]/body[1]/...' → '/html[1]/body[1]/...'
    """
    # 1) 문자열 앞의 '/[document][n]/' 패턴만 없애는 헬퍼
    def _fix(xp: str) -> str:
        return re.sub(r'^\s*/\s*\[document\](?:\[\d+\])?/?', '/', xp)

    # 2) 공통 치환기: (By.XPATH, "...") 캡처해서 내부만 교체
    def _repl_tuple(m):
        quote = m.group(2)
        xp = m.group(3)
        fixed = _fix(xp)
        return f"{m.group(1)}{quote}{fixed}{quote}{m.group(4)}"

    text = g
    # (a) (By.XPATH, "…")  — EC, until 등 모든 튜플 인자
    text = re.sub(r'(\(\s*By\.XPATH\s*,\s*)(["\'])(.+?)\2(\s*\))', _repl_tuple, text, flags=re.DOTALL)
    # (b) .find_element(By.XPATH, "…")
    text = re.sub(r'(\.find_element\(\s*By\.XPATH\s*,\s*)(["\'])(.+?)\2(\s*\))', _repl_tuple, text, flags=re.DOTALL)
    # (c) .find_elements(By.XPATH, "…")
    text = re.sub(r'(\.find_elements\(\s*By\.XPATH\s*,\s*)(["\'])(.+?)\2(\s*\))', _repl_tuple, text, flags=re.DOTALL)

    return text


def _sanitize_locator_tuples(code: str) -> str:
    """
    로케이터 주변의 잘못된 꼬리 토큰을 정리하고 XPath의 '['와 ']' 균형을 맞춘다.
    - 커버 범위:
      1) (By.XPATH, "…")          ← EC.presence_of_element_located((By.XPATH, "…"))
      2) .find_element(By.XPATH, "…")
      3) .find_elements(By.XPATH, "…")
      4) 그 외 By.XPATH, "…" 일반 패턴(제네릭 백업)
    - 동작:
      - 문자열 내부의 '[' 개수와 ']' 개수를 비교해 부족한 ']'를 **문자열 안에 추가**
      - 튜플/호출 괄호 **바깥**에 남은 고아 ']'는 제거
      - 드문 중복 따옴표/괄호 꼬임도 보수 정리
    """
    import re

    def _balance(expr: str, trailing: str) -> tuple[str, str]:
        opens = expr.count("[")
        closes = expr.count("]")
        deficit = max(0, opens - closes)

        # 1) 뒤에 붙은 ']'가 있으면 먼저 흡수
        if deficit > 0 and "]" in trailing:
            borrow = min(deficit, trailing.count("]"))
            if borrow:
                expr += "]" * borrow
                trailing = trailing.replace("]", "", borrow)
                deficit -= borrow

        # 2) 여전히 부족하면 문자열 안에 직접 보충
        if deficit > 0:
            expr += "]" * deficit

        # 3) 남은 외부 ']'는 제거
        trailing = re.sub(r"\s*\]+", "", trailing)
        return expr, trailing

    # ❶ (By.XPATH, "…")
    def _repl_tuple(m: re.Match) -> str:
        prefix, quote, expr, suffix, trailing = m.groups()
        expr, trailing = _balance(expr, trailing or "")
        return f"{prefix}{quote}{expr}{quote}{suffix}{trailing}"

    pat_tuple = re.compile(
        r'(\(By\.XPATH\s*,\s*)(["\'])([\s\S]*?)\2(\))(\s*\]+)?',
        flags=re.IGNORECASE
    )
    code = pat_tuple.sub(_repl_tuple, code)

    # ❷ .find_element(By.XPATH, "…")
    def _repl_find(m: re.Match) -> str:
        prefix, quote, expr, suffix, trailing = m.groups()
        expr, trailing = _balance(expr, trailing or "")
        return f"{prefix}{quote}{expr}{quote}{suffix}{trailing}"

    pat_find = re.compile(
        r'(\.find_element\(\s*By\.XPATH\s*,\s*)(["\'])([\s\S]*?)\2(\s*\))(\s*\]+)?',
        flags=re.IGNORECASE
    )
    code = pat_find.sub(_repl_find, code)

    # ❸ .find_elements(By.XPATH, "…")
    pat_finds = re.compile(
        r'(\.find_elements\(\s*By\.XPATH\s*,\s*)(["\'])([\s\S]*?)\2(\s*\))(\s*\]+)?',
        flags=re.IGNORECASE
    )
    code = pat_finds.sub(_repl_find, code)

    # ❹ 제네릭 백업: By.XPATH, "…" 형태가 남아있을 경우 균형 보정
    def _repl_generic(m: re.Match) -> str:
        prefix, quote, expr = m.groups()
        expr, _ = _balance(expr, "")
        return f"{prefix}{quote}{expr}{quote}"

    pat_generic = re.compile(
        r'(By\.XPATH\s*,\s*)(["\'])([\s\S]*?)\2',
        flags=re.IGNORECASE
    )
    code = pat_generic.sub(_repl_generic, code)

    # ❺ 공통 마무리: 로케이터 튜플 뒤의 고아 ']' 제거 + 중복 따옴표/괄호 보수
    code = re.sub(
        r'(\(By\.(?:ID|XPATH|CSS_SELECTOR|CLASS_NAME)\s*,\s*)(["\'])([\s\S]*?)\2(\))\s*\]',
        r'\1\2\3\2\4',
        code,
        flags=re.IGNORECASE
    )
    code = re.sub(r'"\)\s*"\)', '"))', code)
    code = re.sub(r"'\)\s*'\)", "'))", code)
    return code




def _ensure_compilable_python(code: str) -> str:
    """
    결과 코드의 Python 구문 유효성을 보장. 실패 시 1차 살균을 거쳐 재시도하고,
    그래도 실패하면 SyntaxError를 재전파하여 상위에서 원인 라인을 확인할 수 있게 함.
    """
    import ast
    try:
        ast.parse(code)
        return code
    except SyntaxError:
        fixed = _sanitize_locator_tuples(code)
        try:
            ast.parse(fixed)
            return fixed
        except SyntaxError as e:
            raise SyntaxError(f"[postprocess] syntax fix failed: {e}") from e


def postprocess_generated_code(generated_code: str) -> str:
    """LLM 생성 코드에 대한 표준 후처리 파이프라인
    - 코드펜스 제거 → anti-flaky sleep → 표준 tearDown 교체 → unittest.main 패치
      → 필수 import 보강 → '/[document]' XPath 프리픽스 제거
    - 목적: 실행 안정성과 팀 코딩 규약 준수 확보
    """
    generated_code = remove_markdown(generated_code)
    generated_code = insert_sleep_before_assert(generated_code)
    generated_code = patch_teardown(generated_code)
    generated_code = patch_unittest_main(generated_code)
    # ▼ 생성 코드에서 case_id 계산 시 os 사용 → import 누락 방지
    generated_code = ensure_import(generated_code, "import os")
    generated_code = ensure_import(generated_code, "import unittest")
    generated_code = ensure_import(generated_code, "from selenium.webdriver.common.by import By")
    generated_code = ensure_import(generated_code, "from selenium.webdriver.support.ui import WebDriverWait")
    generated_code = ensure_import(generated_code, "from selenium.webdriver.support import expected_conditions as EC")
    generated_code = ensure_import(generated_code, "from selenium.webdriver.common.keys import Keys")
    generated_code = ensure_import(generated_code, "import default_setting")
    generated_code = strip_document_prefix_xpaths(generated_code)
    return generated_code


# ========================= [URL & DOM FILTERING] =========================
def filter_and_order_docs_by_urls(dom_docs, preferred_paths):
    """입력한 URL/베이스경로에 해당하는 문서만 남기고, 경로 순서대로 앞으로 정렬.
    - 목적: 시나리오 단계의 URL 스코프에 맞는 DOM 후보를 우선 노출
    - 매칭 실패 시 원본 유지(보수적 전략)
    """
    if not preferred_paths:
        return dom_docs
    def belongs(url: str) -> bool:
        return any(p and p in url for p in preferred_paths)
    kept = []
    for d in dom_docs:
        m = getattr(d, "metadata", {}) or {}
        u = (m.get("url") or "")
        if belongs(u):
            kept.append(d)
    if not kept:
        return dom_docs  # 매칭 없으면 안전하게 원본 유지
    def ord_key(d):
        u = (getattr(d, "metadata", {}) or {}).get("url", "")
        for i, p in enumerate(preferred_paths):
            if p and p in u:
                return i
        return len(preferred_paths) + 1
    kept.sort(key=ord_key)
    return kept


def docs_for_base(dom_docs, base_path: str):
    """특정 단계의 base_path에 해당하는 문서만 반환(없으면 빈 목록).
    - 단일 단계 컨텍스트 축소용
    """
    if not base_path:
        return []
    picked = []
    for d in dom_docs:
        m = getattr(d, "metadata", {}) or {}
        u = m.get("url") or ""
        if base_path in u:
            picked.append(d)
    return picked


def parse_step_urls(lines):
    """
    각 단계 문자열에서 http.../vpes/... 또는 /vpes/... URL을 추출.
    반환: (clean_steps, step_base_paths, url_tags_for_prompt)
      - clean_steps: URL 제거 후의 단계 텍스트(LLM에게 행동 문구만 전달)
      - step_base_paths: 각 단계가 참조할 base_path(/vpes/<section>)
         · 해당 줄에 URL이 있으면 그걸 사용, 없으면 None (직전 URL 계승 금지 — 명시적 스코프만 인정)
      - url_tags_for_prompt: 각 단계 뒤에 붙일 '@URL:/vpes/...' 태그 문자열
    구현:
      - URL 정규식으로 라인에서 URL 찾아내고, base_path로 정규화한 뒤 라인에서는 제거
      - 프롬프트 구성 시 '@URL:...' 꼬리표로 재부착(컨텍스트 선별용)
    """
    url_pat = re.compile(r'((?:https?://[^\s,)\]　]+)|(?:/vpes/[^\s,)\]　]+))', re.IGNORECASE)
    clean_steps = []; step_base_paths = []; url_tags = []
    for s_raw in lines:
        m = url_pat.search(s_raw)
        if m:
            u = m.group(1).strip()
            bp = derive_base_path(u) or u
            s2 = (s_raw[:m.start()] + s_raw[m.end():]).strip().rstrip(',').rstrip()
            clean_steps.append(s2)
            step_base_paths.append(bp)
            url_tags.append(f' @URL:{bp}')
        else:
            clean_steps.append(s_raw.strip())
            step_base_paths.append(None)
            url_tags.append('')
    return clean_steps, step_base_paths, url_tags


# ========================= [CHROMA: SELECT/RETRIEVE] =========================
def select_chroma_dirs(base_dir: str, user_inputs: list[str]) -> list[str]:
    """사용자 입력(베이스경로/키워드)와 일치하는 폴더만 선별
    - 목적: RAG 대상 DB를 줄여 검색 비용/잡음 감소
    - 전략: 폴더명 소문자 포함 여부(any)
    """
    if not os.path.isdir(base_dir):
        return []
    keys = [u.strip().lower() for u in user_inputs if u and u.strip()]
    if not keys:
        return []
    selected = []
    for name in os.listdir(base_dir):
        p = os.path.join(base_dir, name)
        if not os.path.isdir(p):
            continue
        low = name.lower()
        if any(k in low for k in keys):
            selected.append(p)
    return selected


def retrieve_from_selected_dirs(
        selected_dirs: list[str],
        query: str,
        embedding,
        k_total: int = 200,
        mmr_fetch_factor: int = 5,
        mmr_lambda: float = 0.35
):
    """
    선택된 여러 DB 폴더 각각에서 MMR 검색을 수행하고 결과를 합칩니다.
    - 각 DB당 균등 할당(per_db)으로 과도한 편중 방지
    - fetch_k는 per_db * mmr_fetch_factor (상한 1000)
    - 실패한 DB는 경고 후 건너뜀(부분내성)
    반환: LangChain Document들의 합치 목록 상위 k_total개
    """
    if not selected_dirs:
        return []
    per_db = max(1, math.ceil(k_total / len(selected_dirs)))
    fetch_k = min(per_db * mmr_fetch_factor, 1000)
    all_docs = []
    for db_dir in selected_dirs:
        try:
            client = chromadb.PersistentClient(path=db_dir)
            vs = Chroma(
                client=client,
                collection_name="dom_elements",
                embedding_function=embedding,
            )
            docs = vs.max_marginal_relevance_search(
                query=query,
                k=per_db,
                fetch_k=fetch_k,
                lambda_mult=mmr_lambda
            )
            all_docs.extend(docs)
        except Exception as e:
            print(f"[경고] DB 접근 실패, 건너뜀: {db_dir} ({e})")
    return all_docs[:k_total]


def retrieve_dom_docs_per_base(
    base_paths: list[str],
    clean_steps: list[str],
    step_base_paths: list[str],
    embedding,
    selected_dirs: list[str],
):
    """베이스 경로별로 쿼리를 구성해 독립적으로 MMR 검색을 수행
    - 각 base_path에 연관된 단계 텍스트만 묶어 query_for_bp로 사용(없으면 bp 자체를 쿼리)
    - selected_dirs 중 base_path의 마지막 세그먼트가 포함된 폴더를 우선 사용, 없으면 전체로 fallback
    - 결과를 모두 합쳐 반환
    """
    from collections import defaultdict
    all_docs = []
    steps_by_base = defaultdict(list)
    for txt, bp in zip(clean_steps, step_base_paths):
        if bp:
            steps_by_base[bp].append(txt)

    for bp in base_paths:
        if not bp:
            continue
        dirs_for_bp = selected_dirs[:]
        if not dirs_for_bp:
            continue
        key_fragment = os.path.basename(bp).lower()
        dirs_for_bp = [d for d in dirs_for_bp if key_fragment in os.path.basename(d).lower()]
        if not dirs_for_bp:
            dirs_for_bp = selected_dirs  # 🔁 fallback
        query_for_bp = "\n".join(steps_by_base.get(bp, [])) or bp
        docs_bp = retrieve_from_selected_dirs(
            selected_dirs=dirs_for_bp,
            embedding=embedding,
            query=query_for_bp,
            k_total=K_PER_BASE,
            mmr_fetch_factor=5,
            mmr_lambda=0.35,
        )
        all_docs.extend(docs_bp)
    return all_docs


def _get_all_docs(coll, batch: int = 500):
    """Chroma 컬렉션에서 전체 document를 페이징으로 수집
    - include=["documents"]로 문서만 가져오고, 간단한 Doc 객체로 변환
    """
    docs, offset = [], 0
    while True:
        got = coll.get(include=["documents"], limit=batch, offset=offset)
        ids = got.get("ids") or []
        if not ids:
            break
        docs.extend(got.get("documents") or [])
        offset += len(ids)
    return [type("Doc", (object,), {"metadata": {}, "page_content": d}) for d in docs]


def safe_load_collection(persist_dir: str, collection_name: str):
    """
    지정 Chroma 디렉토리에서 컬렉션 전체 문서를 페이징으로 로드.
    - 실패해도 예외를 전파하지 않고 빈 리스트 반환(호출부 연속성 유지)
    """
    try:
        client = chromadb.PersistentClient(path=persist_dir)
        coll = client.get_collection(name=collection_name)
        return _get_all_docs(coll, batch=1000)
    except Exception:
        return []


# ========================= [STEP: EXTRACT/MATCH/REF] =========================
def extract_step_comments_from_code(code: str) -> list[str]:
    """생성된 코드 안의 '# N. 설명' 형식 주석을 추출하여 'N. 설명' 리스트로 반환
    - 시나리오-스크립트 적합성 비교를 위한 요약용
    """
    pat = re.compile(r'^\s*#\s*(\d+)\.\s*(.+?)\s*$', re.MULTILINE)
    return [f"{m.group(1)}. {m.group(2).strip()}" for m in pat.finditer(code)]


def normalize_step_title(s: str) -> str:
    """스텝 제목 포맷 정규화
    - '1. 저장 버튼 클릭' → '저장 버튼 클릭'
    - 소문자/다중 공백 정규화
    - 참고코드 매칭의 견고성 제고
    """
    s = re.sub(r'^\s*\d+\s*\.\s*', '', s.strip())
    s = re.sub(r'\s+', ' ', s)
    return s.lower()


def extract_step_blocks_from_code(code: str):
    """참고 코드에서 각 Step 주석과 그 다음 Step 주석 직전까지의 '코드 블록' 추출
    - 반환: [{'num':1,'title':'...','code':'...'}...]
    - 방법: 라인 인덱스를 모두 모아 구간별 슬라이싱
    """
    lines = code.splitlines()
    idxs = []
    pat = re.compile(r'^\s*#\s*(\d+)\.\s*(.+?)\s*$')
    for i, line in enumerate(lines):
        m = pat.match(line)
        if m:
            idxs.append((i, int(m.group(1)), m.group(2).strip()))
    blocks = []
    for j, (start, num, title) in enumerate(idxs):
        end = idxs[j+1][0] if j+1 < len(idxs) else len(lines)
        code_block = "\n".join(lines[start:end]).strip()
        blocks.append({"num": num, "title": title, "code": code_block})
    return blocks


def _cosine(v1, v2) -> float:
    """코사인 유사도(간단 구현). 벡터가 비었을 경우 0"""
    dot = sum(a*b for a,b in zip(v1, v2))
    n1 = math.sqrt(sum(a*a for a in v1))
    n2 = math.sqrt(sum(b*b for b in v2))
    return (dot/(n1*n2)) if (n1 and n2) else 0.0


def text_overlap_score(a: str, b: str) -> float:
    """텍스트 유사도(자카드 + 시퀀스매쳐 평균)
    - 정규화/토큰화 후 겹치는 정도와 편집적 유사도를 결합
    """
    ta = set(re.findall(r'[A-Za-z0-9가-힣]+', a.lower()))
    tb = set(re.findall(r'[A-Za-z0-9가-힣]+', b.lower()))
    jacc = len(ta & tb) / max(1, len(ta | tb))
    seq = SequenceMatcher(None, a.lower(), b.lower()).ratio()
    return (jacc + seq) / 2


def compute_step_embeddings(texts, embedding):
    """스텝 제목 목록에 대한 임베딩 계산
    - OpenAIEmbeddings.embed_query를 개별 호출(간단 구현)
    - 실패 시 빈 벡터 수용(부분내성)
    """
    vecs = []
    for t in texts:
        try:
            vecs.append(embedding.embed_query(t))
        except Exception:
            vecs.append([])
    return vecs


def match_steps_per_title(user_steps: list[str], ref_blocks: list[dict], embedding, sim_threshold: float = 0.62):
    """
    사용자 단계(문구) vs 참고코드 단계(주석 문구) 유사도 매칭.
    - 순서 무관, 1:1 그리디 매칭
    - 스코어: 0.5*임베딩코사인 + 0.3*텍스트겹침 + 0.2*시퀀스유사
    - sim_threshold 이상인 페어만 채택
    반환: {user_index(1-based): {'ref_num':..., 'score':..., 'ref_title':..., 'code':...}}
    """
    user_norms = [normalize_step_title(s) for s in user_steps]
    ref_norms  = [normalize_step_title(b['title']) for b in ref_blocks]

    # 임베딩 미리 계산
    u_vecs = compute_step_embeddings(user_norms, embedding)
    r_vecs = compute_step_embeddings(ref_norms, embedding)

    # 모든 쌍 점수 계산
    pairs = []
    for ui, ut in enumerate(user_norms):
        for ri, rt in enumerate(ref_norms):
            emb = _cosine(u_vecs[ui], r_vecs[ri]) if (u_vecs[ui] and r_vecs[ri]) else 0.0
            overl = text_overlap_score(ut, rt)
            seq_only = SequenceMatcher(None, ut, rt).ratio()
            score = 0.5*emb + 0.3*overl + 0.2*seq_only
            pairs.append((score, ui, ri))

    # 점수 높은 순으로 그리디 매칭
    pairs.sort(reverse=True, key=lambda x: x[0])
    used_u, used_r = set(), set()
    mapping = {}
    for score, ui, ri in pairs:
        if score < sim_threshold:
            break
        if ui in used_u or ri in used_r:
            continue
        used_u.add(ui); used_r.add(ri)
        b = ref_blocks[ri]
        mapping[ui+1] = {
            "ref_num": b["num"],
            "score": round(score, 3),
            "ref_title": b["title"],
            "code": b["code"],
        }
    return mapping


def build_per_step_reference_context(user_steps: list[str], step_mapping: dict, base_paths: list[str]) -> str:
    """
    LLM 프롬프트에 넣을 '단계별 참고 코드' 텍스트를 구성
    - 각 사용자 단계에 대해 매칭된 참고 코드를 CODE 블록으로 제공
    - URL 힌트([REF_BASE_PATH_HINT])는 선택적으로 제공
    """
    parts = []
    if base_paths:
        parts.append("[REF_BASE_PATH_HINT]\n" + ", ".join(base_paths[:8]))
    for i, step in enumerate(user_steps, start=1):
        if i in step_mapping:
            m = step_mapping[i]
            parts.append(
                f"[REF_STEP {i}] matched_from={m['ref_num']} score={m['score']}\n"
                f"TITLE: {m['ref_title']}\n"
                "CODE:\n```python\n" + m["code"] + "\n```"
            )
    return "\n\n".join(parts)


# ========================= [HINT-BASED SELECTOR PATCH] =========================
def replace_hint_per_step(code: str, step_hints_map: dict[int, dict[str, tuple[str, str]]]) -> str:
    """
    대괄호 힌트 기반 선택자 치환을 '스텝별 코드 블록'에 한정하여 수행
    - step_hints_map: { step_idx: { hint_lower: ("ID"|"XPATH", value), ... }, ... }+
    - 대상: 각 스텝(# N. ...) 주석부터 다음 스텝 직전까지
    - 원칙: By.ID로 이미 선택된 로케이터는 유지, By.XPATH만 힌트로 교체(조건 충족 시)
    구현:
      - 정규식으로 스텝별로 코드 블록을 분할
      - 해당 스텝의 힌트 사전을 찾고, XPATH 튜플 문자열을 검사하여 힌트 키워드 포함 시 치환
    """
    parts = re.split(r'(^\s*#\s*\d+\.\s*.*$)', code, flags=re.MULTILINE)
    out, curr_step = [], None

    # 개선된 패턴: 여는 따옴표와 닫는 따옴표를 동일 그룹(백레퍼런스 \2)로 강제, 내용은 DOTALL로 캡처
    PAT_XPATH_TUPLE = re.compile(
        r'(\(\s*By\.XPATH\s*,\s*)(["\'])(.*?)\2(\s*\))',
        flags=re.IGNORECASE | re.DOTALL
    )

    for chunk in parts:
        m = re.match(r'^\s*#\s*(\d+)\.\s*', chunk)
        if m:
            curr_step = int(m.group(1))
            out.append(chunk)
            continue

        if curr_step and curr_step in step_hints_map:
            repl_map = step_hints_map[curr_step]

            def _sub_xpath(mo: re.Match) -> str:
                xpath = mo.group(3)  # 내용만
                xlow = xpath.lower()
                for hint_lower, (typ, val) in repl_map.items():
                    if hint_lower in xlow:
                        if typ == "ID":
                            return f'(By.ID, "{val}")'
                        else:
                            return f'(By.XPATH, "{val}")'
                return mo.group(0)

            # By.XPATH 만 대상으로, 힌트 단어가 포함된 경우에만 치환
            chunk = PAT_XPATH_TUPLE.sub(_sub_xpath, chunk)

        out.append(chunk)

    return ''.join(out)


def convert_bracket_hints_to_exact_selectors_per_step(generated_code: str, dom_docs, question: str) -> str:
    """
    '[힌트]'가 포함된 시나리오 각 단계에 대해, 해당 스텝의 @URL 스코프 DOM 안에서
    힌트와 일치하는 요소(ID 우선, 없으면 XPATH)를 찾아 선택자 치환(스텝 블록 범위 내).
    - question: '1. ... @URL:/vpes/SectionA\n2. ... @URL:/vpes/SectionB ...' 형태의 문장
    - 흐름:
      1) question을 줄단위로 파싱 → (스텝번호, 텍스트, @URL 스코프) 추출
      2) 스코프에 맞는 DOM 문서 집합을 선별
      3) 각 힌트에 대해 (태그일치 → desc/class/xpath 포함) 우선순위로 가장 좋은 로케이터 선택
      4) 스텝별 힌트 맵(step_hints_map)을 구성
      5) replace_hint_per_step로 해당 스텝 코드 블록에만 치환 적용
    """
    step_hints_map: dict[int, dict[str, tuple[str, str]]] = {}
    lines = [ln for ln in (question or "").splitlines() if ln.strip()]

    for ln in lines:
        m_num = re.match(r'^\s*(\d+)\.\s*(.*)$', ln)
        if not m_num:
            continue
        step_idx = int(m_num.group(1))
        text = m_num.group(2)

        # 힌트와 base_path(@URL:...) 추출
        hints = re.findall(r'\[([^\]]+)\]', text)
        m_url = re.search(r'@URL:([^\s]+)', ln)
        base_path = m_url.group(1).strip() if m_url else None

        if not hints:
            continue

        # 스코프 문서 선택
        scoped = docs_for_base(dom_docs, base_path) if base_path else dom_docs

        for hint in hints:
            hint_lower = hint.lower()
            best = None

            # 1) 태그 이름 우선
            if hint_lower in ['span', 'div', 'button', 'input', 'a', 'li', 'p', 'h1', 'h2', 'h3']:
                for doc in scoped:
                    meta = getattr(doc, "metadata", {}) or {}
                    if (meta.get("tag") or "").lower() == hint_lower:
                        if meta.get("id"):
                            best = ("ID", meta["id"]); break
                        if meta.get("xpath"):
                            best = ("XPATH", meta["xpath"]); break

            # 2) desc/class/xpath 포함 매칭
            if not best:
                for doc in scoped:
                    meta = getattr(doc, "metadata", {}) or {}
                    desc = (meta.get("desc") or "").lower()
                    clss = (meta.get("class") or "").lower()
                    xpth = (meta.get("xpath") or "").lower()
                    if (hint_lower in desc) or (hint_lower in clss) or (hint_lower in xpth):
                        if meta.get("id"):
                            best = ("ID", meta["id"]); break
                        if meta.get("xpath"):
                            best = ("XPATH", meta["xpath"]); break

            if best:
                step_hints_map.setdefault(step_idx, {})[hint_lower] = best

    # 스텝별로만 치환 적용
    return replace_hint_per_step(generated_code, step_hints_map)


# ========================= [RUNTIME: EXECUTION & LOGGING] =========================
def _script_dir() -> str:
    """현재 스크립트 파일의 디렉토리 반환
    - __file__이 없는 환경(인터프리터)에서는 CWD 사용
    """
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except NameError:
        return os.getcwd()


def run_generated_test(file_path: str, timeout_sec: int = 900,
                       extra_args: list[str] | None = None,
                       extra_env: dict[str, str] | None = None) -> dict:
    """
    생성된 unittest 스크립트를 별도 프로세스로 실행하고 결과 요약 딕셔너리를 반환.
    - 목적: 테스트 실행을 메인 프로세스와 분리, 타임아웃/인코딩/환경 격리
    - 구현:
      * PYTHONPATH에 스크립트 폴더를 선두 추가(로컬 모듈 import 보장)
      * stdout/stderr 바이너리를 UTF-8 우선 디코딩, 실패 시 cp949 폴백
      * 결과 상태 PASS/FAIL/UNKNOWN/TIMEOUT 판정 및 tail 보관
    """
    import copy
    script_dir = os.path.dirname(os.path.abspath(file_path))
    cmd = [sys.executable, file_path] + (extra_args or [])
    env = copy.deepcopy(os.environ)
    if extra_env:
        env.update(extra_env)

    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("PYTHONUTF8", "1")
    # 🔧 스크립트 폴더를 PYTHONPATH 맨 앞에 추가
    sep = os.pathsep
    env["PYTHONPATH"] = script_dir + (sep + env["PYTHONPATH"] if "PYTHONPATH" in env and env["PYTHONPATH"] else "")

    start = time.time()
    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=False,
            timeout=timeout_sec,
            env=env,
            cwd=script_dir,  # 🔧 작업 디렉터리 고정
        )
        duration = round(time.time() - start, 2)
        def _decode(b: bytes) -> str:
            try:
                return b.decode("utf-8")  # 1차 시도
            except UnicodeDecodeError:
                # 로캘 미가용 환경 대비 간단 폴백
                return b.decode("cp949", errors="replace")

        out = _decode(proc.stdout or b"")
        err = _decode(proc.stderr or b"")
        if re.search(r'FAILED\s*\(', out, re.IGNORECASE) or "Traceback (most recent call last):" in out or proc.returncode != 0:
            status = "FAIL"
        elif re.search(r'^\s*OK\s*$', out, re.MULTILINE):
            status = "PASS"
        else:
            status = "UNKNOWN"
        return {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "case_id": os.path.splitext(os.path.basename(file_path))[0],
            "status": status,
            "duration_sec": duration,
            "returncode": proc.returncode,
            "file": file_path,
            "stdout_tail": out[-20000:],
            "stderr_tail": err[-20000:],
        }
    except subprocess.TimeoutExpired as e:
        duration = round(time.time() - start, 2)
        return {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "case_id": os.path.splitext(os.path.basename(file_path))[0],
            "status": "TIMEOUT",
            "duration_sec": duration,
            "returncode": -1,
            "file": file_path,
            "stdout_tail": (getattr(e, "stdout", "") or "")[-20000:],
            "stderr_tail": (getattr(e, "stderr", "") or "")[-20000:],
        }


def append_result_to_excel(xlsx_path: str, result: dict) -> None:
    """
    실행 결과를 test_result.xlsx에 누적 기록
    - 파일이 없으면 생성 후 헤더 작성
    - 있으면 마지막 행 뒤에 append
    - 헤더 컬럼: timestamp, case_id, status, duration_sec, returncode, file, stdout_tail, stderr_tail
    """
    from openpyxl import Workbook, load_workbook

    headers = ["timestamp", "case_id", "status", "duration_sec", "returncode", "file", "stdout_tail", "stderr_tail"]
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # 헤더가 없으면 생성
        if ws.max_row == 0:
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws = wb.active
    ws.append([result.get(h, "") for h in headers])
    wb.save(xlsx_path)


def run_and_log(file_name: str, cli_args: list[str] | None = None,
                env_vars: dict[str, str] | None = None) -> dict:
    """
    테스트 파일을 실행하고 결과를 엑셀에 누적 저장하는 편의 래퍼
    - 내부적으로 run_generated_test → append_result_to_excel 실행
    - 콘솔에 요약 메세지 출력
    """
    base_dir = _script_dir()
    test_file = os.path.join(base_dir, file_name)
    result = run_generated_test(test_file, timeout_sec=900,
                                extra_args=cli_args, extra_env=env_vars)
    xlsx_path = os.path.join(base_dir, "test_result.xlsx")
    append_result_to_excel(xlsx_path, result)
    print(f"\n🗂️ 자동 실행 결과: {result['status']} (returncode={result['returncode']}, duration={result['duration_sec']}s)")
    print(f"{result['stderr_tail']}")
    print(f"➡️ 결과가 '{xlsx_path}'에 누적 저장되었습니다.")
    return result


# ========================= [MAIN PIPELINE] =========================
def main():
    """
    전체 파이프라인 오케스트레이션
    1) 환경변수 로딩 및 LLM/임베딩 초기화
    2) 사용자 입력 수집(테스트케이스명, 메뉴 섹션, 제목, 단계별 시나리오)
    3) 단계별 URL 파싱 → base_paths 산출 → 대상 Chroma DB 선택
    4) (선택) 유사 시나리오 파일의 Step 주석과 사용자 단계 매칭 → per-step reference 구성
    5) base_path별 RAG(MMR) 검색으로 DOM 후보 집합 수집/정렬/중복제거
    6) default_setting/move_menu 함수 요약 로드
    7) LLM 프롬프트 구성(context/function_context/step_reference_context/question)
    8) LLM 호출 → 코드 생성 → 클래스/메서드명 치환
    9) 커스텀 검증·보정(validate_generated_code, 힌트 기반 치환, 표준 후처리)
    10) 코드 파일 저장 및 시나리오-스크립트 적합성 평가(JSON)
    11) 통과 시 샘플 인자/환경으로 자동 실행 및 로깅
    """
    # 1) 환경 변수 로딩
    load_dotenv()
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        raise RuntimeError("OPENAI_API_KEY 환경변수가 설정되지 않았습니다. .env 또는 환경변수를 확인하세요.")

    # 2) LLM/임베딩 초기화
    #llm = ChatOpenAI(model="gpt-5")
    llm = ChatOpenAI(
        model="gpt-5-chat-latest",
        use_responses_api=True,  # Responses API 사용 권장
        output_version="responses/v1",  # reasoning summary·내장툴 대응 포맷
    )
    embedding = OpenAIEmbeddings(model="text-embedding-3-large")

    # 4) Prompt 정의 — per-step reference 적용
    prompt = PromptTemplate(
        input_variables=["context", "question", "function_context", "step_reference_context"],
        template="""
# ===== CoT 적용 버전 (출력은 코드만) =====
단계적으로 내부에서 생각하되, 최종 출력에는 절대 사고 과정을 포함하지 마세요.
당신은 Selenium 테스트 자동화 코드를 생성하는 전문가입니다.

아래는 웹 페이지의 HTML DOM 요소 목록(context),
기능 함수 요약(function_context),
테스트 시나리오(question),
그리고 **각 단계별 참고 코드(step_reference_context)** 입니다.

# [내부 사고 프로세스 — 출력 금지]
- 내부적으로만 다음 단계를 거쳐 최선의 답을 선택하고, 최종 결과(테스트 코드)만 출력할 것.
  1) 시나리오 파싱 → 단계별 행동으로 분해.
  2) 각 **단계 라인 끝의 '@URL:/vpes/…' 태그**가 있으면 해당 URL 범위의 요소만 사용.
     - 태그가 없으면 DOM context를 사용하지 말고 function_context 함수만 사용.
  3) 함수 재사용 매핑 → function_context를 스캔해 동일/유사 기능 우선 매핑.
  4) **per-step 참고 코드 적용 규칙**
     - step_reference_context에 [REF_STEP i]가 있으면, 현재 i단계 구현 시 해당 CODE를 **가능한 한 그대로 재사용**.
     - 단, @URL 스코프가 다르거나 의미가 어긋나면 재사용 금지.
     - By.ID로 지정된 셀렉터는 그대로 유지(변경 금지).
     - 변수명/헬퍼 호출 정도의 최소 수정만 허용.
  5) DOM 매핑 → context의 text/description/aria/placeholder와 의미적으로 정합되게 선택.
  6) 중복 이동 제거: login 후 /vpes 재이동 금지 등.
  7) 코드 설계: unittest.TestCase 골격 → setUp → TC 메서드 → tearDown → main 순서.
  8) 주석 규칙: 각 주요 단계에 “숫자. 한 줄 요약”만.
  9) 형식 점검: import 최상단, case_id 추출, URL 조립 규칙, 선택자 규칙, 함수 재사용 준수.
 10) 자기검토: 불필요한 send_keys()/click()/driver.get() 제거.
 11) 최종 산출: 아래 [최종 출력 규칙]을 지켜 오직 코드블록만 출력.

# [기본 테스트 코드 구조]
- unittest.TestCase 사용.
- setUp()은 반드시 default_setting.setup()으로 WebDriver 초기화.
- case_id는 import 아래:
  => case_id = os.path.basename(os.path.splitext(__file__)[0])
- 각 주요 단계에 **숫자. 한 줄 요약** 주석.
- tearDown()은 테스트 메서드 다음.
- if __name__ == "__main__": 포함.
- import 선언은 모두 맨 윗단.

# [기존 함수 활용 지침]
- function_context는 default_setting.py 및 move_menu.py의 요약.
- default_setting.login(driver)는 로그인 후 /vpes로 이동. 이후 중복 이동 금지.
- driver.get()이 포함된 함수는 중복 호출 피하기.
- 제공된 함수가 시나리오와 의미가 같다면 반드시 재사용.

# [추가 규칙 — 앵커(직전 클릭 요소) 근접도]
- 앵커: 직전에 클릭/체크한 요소의 XPath(또는 ID→context의 XPath).
- 저장/닫기/확인/적용 같은 버튼을 고를 땐, 같은 @URL 컨텍스트에서 후보를 모으고 아래 우선순위로 선택:
  1) 앵커 XPath와의 공통 prefix 깊이가 가장 큰 것(= 같은 모달 트리 우선).
  2) modal/footer 관련 class( modal, modal-footer, btn-primary/btn-outline-primary )가 있는 후보 가점.
  3) 동일 부모 div 아래 ‘닫기’와 ‘저장’이 쌍으로 있는 세트 우선.
  4) 동률이면 더 짧은 XPath.
- By.ID로 이미 고른 로케이터는 절대 바꾸지 말 것. //body 폴백 금지. 새 ID/클래스/XPath 생성 금지.

# [셀렉터 규칙]  
- 기본 우선순위:
  1. ID가 있으면 → 반드시 By.ID 사용
  2. ID가 없고 class가 있으면 → By.CLASS_NAME 또는 By.CSS_SELECTOR 사용
  3. 둘 다 없으면 → context의 XPath 사용
- 단, 단계 텍스트에 **대괄호 힌트([…])가 있는 경우에만**, 그 힌트와 일치하는 요소를 최우선으로 선택한다.
  - 이때도 가능한 경우 ID를 먼저 쓰고, ID가 없으면 XPath를 사용한다.

# [ASSERT 규칙]
- page_source 금지. 요소 상태/속성/URL/토스트 등으로 검증.

---

## DOM 요소 목록 (context):
{context}

## 기능 함수 요약 (function_context):
{function_context}

## 단계별 참고 코드 (step_reference_context):
{step_reference_context}

## 테스트 시나리오 (question):
{question}
---

# [최종 출력 규칙]
- 오직 **실행 가능한 Python 코드**만 하나의 코드블록으로 출력: ```python ... ```
- 내부 사고, 근거, 설명 텍스트 출력 금지(주석 규칙 외).
- tc id 사용하는 class와 함수 외 추가 함수 생성 금지
        """.strip()
    )

    # 5) 테스트케이스명/시나리오 입력
    tc_name = input("테스트케이스명을 입력하세요 (예: C8270): ").strip()

    # === 추가 입력: 메뉴트리 섹션 / TC 제목 (시나리오 입력 전) ===
    menu_tree_section = input("메뉴트리 섹션을 입력하세요: ").strip()
    tc_title = input("TC 제목을 입력하세요: ").strip()

    print("💬 테스트 시나리오를 단계별로 입력하세요 (한 줄에 한 단계).")
    print("    ↳ 각 단계 끝에 /vpes/... 또는 http(s)://.../vpes/... 를 붙이면 그 URL 스코프로 셀렉터를 고릅니다.")
    print("    ↳ 빈 줄(Enter)을 입력하면 종료됩니다.")

    # 표준 입력을 통해 단계 라인 수집
    steps_raw = []
    while True:
        try:
            line = input()
        except EOFError:
            break
        if not line.strip():
            break
        steps_raw.append(line.strip())

    if not steps_raw:
        raise ValueError("최소 1개 이상의 단계가 필요합니다. 예: '1. VPES 로그인'")

    # 5-1) 단계에서 URL 추출 → clean_steps(문구), step_base_paths(스코프), url_tags(@URL:...) 생성
    clean_steps, step_base_paths, url_tags = parse_step_urls(steps_raw)

    # 프롬프트용 질문 문자열(각 단계 뒤에 URL 태그 추가)
    question_lines = [f"{i+1}. {txt}{url_tags[i]}" for i, txt in enumerate(clean_steps)]
    query = "\n".join(question_lines)

    # 6) URL(또는 키워드) 입력 → (단계에 URL이 있으면 생략)
    inputs = [bp for bp in step_base_paths if bp]  # 단계에서 추출한 base_path들을 우선 사용
    if not inputs:
        raw_urls = input('비교할 URL을 추가로 입력하세요 (여러 개면 쉼표로 구분, 없으면 Enter): ').strip()
        inputs = [u.strip() for u in raw_urls.split(",") if u.strip()]
        inputs = [derive_base_path(u) for u in inputs]

    base_paths = list(dict.fromkeys(inputs))
    print("\n🌐 입력/추출된 URL(base) 목록:")
    for u in base_paths:
        print(" -", u)

    # 7) 선택된 DB 폴더만 대상으로 Retrieval
    match_keys = base_paths + [os.path.basename(p) for p in base_paths]
    selected_dirs = select_chroma_dirs(CHROMA_BASE_DIR, match_keys)
    if not selected_dirs:
        print(f"[경고] 입력 텍스트와 일치하는 DB 폴더가 없습니다. 전체 DB 대상으로 검색합니다.")
        selected_dirs = [
            os.path.join(CHROMA_BASE_DIR, d)
            for d in os.listdir(CHROMA_BASE_DIR)
            if os.path.isdir(os.path.join(CHROMA_BASE_DIR, d))
        ]
        if not selected_dirs:
            raise RuntimeError(f"검색 가능한 DB 폴더가 없습니다: {CHROMA_BASE_DIR}")

    print("\n🔎 자동 매칭된 DB 폴더:")
    for d in selected_dirs:
        print(" -", d)

    # ✅ 사용자가 원하는 폴더만 입력 (쉼표 구분)
    raw_selected = input("\n👉 실제 사용할 DB 폴더만 골라 붙여넣으세요 (쉼표 구분, 없으면 자동 선택 사용): ").strip()

    # 사용자가 수동으로 고른 경로가 있으면 그것만 사용(상대/절대 혼합 허용, 정상화)
    if raw_selected:
        selected_dirs = [
            os.path.normpath(s.strip().strip('"').strip("'"))
            for s in raw_selected.split(",") if s.strip()
        ]
        selected_dirs = [
            s if os.path.isabs(s) else os.path.normpath(
                os.path.join(CHROMA_BASE_DIR, os.path.basename(s)) if not s.startswith(
                    "." + os.sep) else os.path.normpath(os.path.join(".", s)))
            for s in selected_dirs
        ]
        selected_dirs = [s for s in selected_dirs if os.path.isdir(s)]
    # else: 분기 제거 — 자동 매칭된 selected_dirs 그대로 사용 (auto_dirs 미정의 버그 제거)

    # ===== (신규) 유사 시나리오 참고 코드 입력 & 단계별 매칭 =====
    ref_path = input("\n📄 유사 시나리오 Python 테스트 파일 경로(없으면 Enter): ").strip()
    step_reference_context = ""
    if ref_path:
        try:
            with open(ref_path, "r", encoding="utf-8") as rf:
                ref_code = rf.read()
            ref_blocks = extract_step_blocks_from_code(ref_code)
            # 사용자 스텝의 '제목'(숫자 제거한 텍스트) 목록
            user_step_titles = [normalize_step_title(s) for s in clean_steps]
            # 매칭 수행
            mapping = match_steps_per_title(user_step_titles, ref_blocks, embedding, sim_threshold=0.62)
            if mapping:
                step_reference_context = build_per_step_reference_context(clean_steps, mapping, base_paths)
                print(f"🔗 단계별 참고 코드 적용: {len(mapping)}개 단계 매칭됨 → 프롬프트에 반영합니다.")
            else:
                print("[참고 보류] 유사 단계 매칭을 찾지 못했습니다. 주석 문구를 더 가깝게 해보세요.")
        except Exception as e:
            print(f"[경고] 참고 코드 로드/분석 실패: {e}")

    # URL별로 완전 분리해서 독립 검색 (URL마다 K_PER_BASE개까지)
    dom_docs_lc = retrieve_dom_docs_per_base(
        base_paths=base_paths,
        clean_steps=clean_steps,
        step_base_paths=step_base_paths,
        embedding=embedding,
        selected_dirs=selected_dirs,
    )

    # LangChain Document → 간단 객체로 변환 (metadata/page_content 인터페이스만 사용)
    dom_docs = [
        type("Doc", (object,), {"metadata": d.metadata, "page_content": d.page_content})
        for d in dom_docs_lc
    ]

    # 중복 제거 (url, xpath, id 기준) — 동일 요소 중복 노출 방지
    seen, dedup = set(), []
    for d in dom_docs:
        m = getattr(d, "metadata", {}) or {}
        key = (m.get("url") or "", m.get("xpath") or "", m.get("id") or "")
        if key in seen:
            continue
        seen.add(key)
        dedup.append(d)
    dom_docs = dedup

    # LangChain Document → function_context 로드(default_setting/move_menu 컬렉션의 문서)
    default_docs = safe_load_collection("./chroma_default_setting", "default_setting")
    move_menu_docs = safe_load_collection("./chroma_move_menu", "move_menu")

    # ===== 핵심: URL 기반으로 DOM 문서를 먼저 필터/정렬
    preferred_paths = base_paths[:]  # 이미 base 형태(/vpes/Section)
    dom_docs = filter_and_order_docs_by_urls(dom_docs, preferred_paths)

    # 디버그: 단계별 base_path와 DOM 수 프린트(스코프 매칭 상태 점검)
    print("\n🧭 단계별 URL 매핑(미리보기):")
    for i, bp in enumerate(step_base_paths, start=1):
        docs_n = len(docs_for_base(dom_docs, bp))
        print(f" - STEP {i:02d}: base_path={bp or '-'} dom_count={docs_n}")

    # 9) context 구성 — LLM에 전달할 DOM 요약 문자열 생성
    def _fmt(meta, key, maxlen=180):
        v = (meta.get(key) or "")
        sv = str(v)
        return (sv[:maxlen] + "…") if len(sv) > maxlen else sv

    BASE_URL_CONST = "http://localhost:38080"  # 클로저 캡쳐 안정화

    sections = []
    for i, (txt, bp) in enumerate(zip(clean_steps, step_base_paths), start=1):
        step_docs = docs_for_base(dom_docs, bp)

        # 대괄호 힌트가 있는 경우, 힌트와 매칭되는 소스파일 그룹을 우선 선택(컨텍스트 품질 향상)
        bracket_hints = re.findall(r'\[([^\]]+)\]', txt)
        if bracket_hints:
            print(f"🎯 STEP {i}에서 대괄호 힌트 발견: {bracket_hints}")

            step_keywords = extract_keywords(txt)

            best_score = 0
            best_docs = step_docs

            # 후보 source_file 분할 후 키워드 매칭 스코어링(도메인 힌트 강화)
            source_files = set()
            for doc in step_docs:
                meta = getattr(doc, "metadata", {}) or {}
                sf = meta.get("source_file", "")
                if sf: source_files.add(sf)

            for source_file in source_files:
                score = 0
                source_lower = source_file.lower()
                txt_lower = txt.lower()
                for keyword in step_keywords:
                    if keyword in source_lower:
                        score += 3
                if "모달" in txt_lower or "창" in txt_lower:
                    if "modal" in source_lower or "popup" in source_lower:
                        score += 2
                if "dropdown" in txt_lower or "더보기" in txt_lower:
                    if "dropdown" in source_lower or "kebab" in source_lower:
                        score += 2
                if "선택" in txt_lower or "select" in txt_lower:
                    if "select" in source_lower or "choice" in source_lower:
                        score += 2
                if score > best_score:
                    candidate_docs = [doc for doc in step_docs
                                      if (getattr(doc, "metadata", {}) or {}).get("source_file", "") == source_file]
                    if candidate_docs:
                        best_score = score
                        best_docs = candidate_docs
                        print(f"   🎯 Context 최적 매칭: {source_file} (점수: {score}, DOM: {len(candidate_docs)}개)")

            filtered_docs = best_docs

            # 1차로 힌트와 일치하는 요소들을 앞쪽에 배치(우선순위 상승)
            prioritized_docs = []
            remaining_docs = []
            for doc in (filtered_docs if filtered_docs else step_docs):
                meta = getattr(doc, 'metadata', {})
                desc = (meta.get("desc") or "").lower()
                tag = (meta.get("tag") or "").lower()
                matched = False
                for hint in bracket_hints:
                    hint_lower = hint.lower()
                    if ((hint_lower in ['span', 'div', 'button', 'input', 'a', 'li'] and tag == hint_lower) or
                            (hint_lower in desc) or
                            (hint_lower in (meta.get("class") or "").lower()) or
                            (hint_lower in (meta.get("xpath") or "").lower())):
                        prioritized_docs.append(doc)
                        matched = True
                        print(f"   ✅ 힌트 '{hint}'와 매칭: {meta.get('xpath', 'N/A')}")
                        print(f"      매칭 상세 - desc: '{desc[:30]}...', xpath: '{(meta.get('xpath') or '')[:50]}...'")
                        break
                if not matched:
                    remaining_docs.append(doc)
            step_docs = prioritized_docs + remaining_docs

        # DOM 후보를 한 줄 요약으로 직조하여 LLM context에 삽입
        frag = "\n".join([
            f"FullURL:{urljoin(BASE_URL_CONST, getattr(doc, 'metadata', {}).get('url',''))} "
            f"ID:{_fmt(getattr(doc, 'metadata', {}),'id')} "
            f"XPATH:{_fmt(getattr(doc, 'metadata', {}),'xpath')} "
            f"TEXT:{_fmt(getattr(doc, 'metadata', {}),'text', 160)} "
            f"DESC:{_fmt(getattr(doc, 'metadata', {}),'desc', 160)} "
            f"ARIA:{_fmt(getattr(doc, 'metadata', {}),'aria-label', 160)} "
            f"PLACEHOLDER:{_fmt(getattr(doc, 'metadata', {}),'placeholder', 160)}"
            for doc in step_docs
        ])
        sections.append(f"[STEP {i} @URL:{bp or '-'}] {txt}\n{frag}")

    # LLM에 넘길 최종 context(문자열)
    context = "\n\n".join(sections)
    function_context = "\n\n".join(doc.page_content[:2000] for doc in (default_docs + move_menu_docs))

    # 참고: 컨텍스트 확장성을 위해 수집된 ID 일부를 로그로 노출(디버깅)
    print("\n📌 context에 포함된 ID 목록(최대 60개 표시):")
    shown = 0
    for doc in dom_docs:
        mid = getattr(doc, "metadata", {}).get("id")
        if mid:
            print(f"- {mid}")
            shown += 1
            if shown >= 60:
                print("... (생략)")
                break

    # 10) LLMChain 실행 — per-step 참고 컨텍스트 포함
    chain = (
        RunnableMap({
            "context": lambda _: context,
            "function_context": lambda _: function_context,
            "question": lambda _: query,
            "step_reference_context": lambda _: step_reference_context
        })
        | prompt
        | llm
        | StrOutputParser()
    )
    generated_code = chain.invoke({})

    # 11) 클래스명/함수명 사용자 입력 값으로 변경(가독/트래킹)
    generated_code = re.sub(r'class [a-zA-Z0-9_]*\(unittest\.TestCase\):', f'class {tc_name}(unittest.TestCase):', generated_code, count=1)
    generated_code = re.sub(r'def test_[a-zA-Z0-9_]*\(', f'def test_{tc_name}(', generated_code, count=1)

    # 12) 후처리/검증/가드(간단 버전 유지)
    generated_code = validate_generated_code(generated_code, dom_docs + default_docs, auto_fix=True)

    # 12-1) 대괄호 힌트 기반 보정 (fallback 용) — 스텝별 범위 내에서만 치환
    generated_code = convert_bracket_hints_to_exact_selectors_per_step(generated_code, dom_docs, query)

    # 치환 이후 살균 1차 적용 (남아있는 '] 등 꼬리 제거 + XPath 균형 복구)
    #generated_code = _sanitize_locator_tuples(generated_code)

    # 정형 후처리 파이프라인(임포트/tearDown/main/sleep/XPath 보정)
    generated_code = postprocess_generated_code(generated_code)

    # 최종 AST 컴파일 보증(실패 시 살균 재시도 후 에러 표시)
    generated_code = _ensure_compilable_python(generated_code)

    # === 추가: 생성 스크립트 최상단에 메뉴트리/요약 주석 삽입(메타정보)
    header_comment = f'# MenuTree : {menu_tree_section}\n# Summary : {tc_title}\n'
    generated_code = header_comment + generated_code

    # 치환 이후 살균 재적용 (남아있는 '] 등 꼬리 제거 + XPath 균형 복구)
    generated_code = _sanitize_locator_tuples(generated_code)

    # 13) 출력 및 저장
    print("\n💻 생성된 Selenium 테스트 코드:")
    print("=" * 60)
    print(generated_code)
    print("=" * 60)

    file_name = f"{tc_name}.py"
    with open(file_name, "w", encoding="utf-8") as f:
        f.write(generated_code)
    print(f"\n✅ 코드가 '{file_name}' 파일로 저장되었습니다.")

    # 14) 시나리오-스크립트 적합성 검토
    # - 스크립트 내 Step 주석을 추출하여 시나리오와 비교
    # - 금지 규칙: Step 주석에 스코프 문자열(@URL:/vpes/..., /vpes/..., http(s)://.../vpes/...) 등장 금지
    code_step_lines = extract_step_comments_from_code(generated_code)
    scenario_text = query
    code_steps_text = "\n".join(code_step_lines) if code_step_lines else "(no step comments found)"

    eval_prompt = PromptTemplate(
        input_variables=["scenario", "code_steps"],
        template=(
            "다음은 사용자가 입력한 테스트 시나리오와 생성된 자동화 스크립트의 Step 주석 목록입니다.\n"
            "당신의 작업: 시나리오의 각 단계가 스크립트 주석에 의해 정확히 구현되었는지 검토하세요.\n\n"
            "검토 기준:\n"
            "- 단계의 의미/핵심 동작 일치 여부 (문구가 조금 달라도 의미가 같으면 OK)\n"
            "- 누락 단계, 순서 오류, 잘못된 URL 스코프 추정, 동작/셀렉터 불일치는 ISSUE\n"
            "- **금지 규칙(필수)**: 스크립트의 Step 주석에는 '시나리오에서 명시된 스코프'가 포함되면 안 됩니다. "
            "스코프는 아래 패턴과 일치하는 문자열을 의미합니다(대소문자 무시):\n"
            "  • '@URL:'로 시작하는 토큰(예: '@URL:/vpes/...')\n"
            "  • '/vpes/'로 시작하는 경로(예: '/vpes/Section...')\n"
            "  • 'http(s)://…/vpes/…' 형태의 전체 URL\n"
            "절차: [시나리오]에서 위 스코프 토큰들을 모두 추출해 집합 S를 만들고, [스크립트_주석] 안에 S의 어떤 항목이라도 "
            "등장하면 각각 ISSUE로 기록하세요. 하나라도 발견되면 compliance=false 입니다.\n\n"
            "출력 형식(JSON만 출력):\n"
            "{{\"compliance\": true|false, \"issues\": [\"사유1\", \"사유2\", \"...\"]}}\n\n"
            "[시나리오]\n{scenario}\n\n[스크립트_주석]\n{code_steps}\n"
        ),
    )

    # 평가 LLM은 JSON 모드 권장(코드펜스 섞임 방지)
    eval_llm = llm.bind(response_format={"type": "json_object"})

    eval_chain = (
        RunnableMap({
            "scenario": lambda _: scenario_text,
            "code_steps": lambda _: code_steps_text
        })
        | eval_prompt
        | eval_llm
        | StrOutputParser()
    )

    eval_raw = eval_chain.invoke({})

    print("\n🧪 시나리오-스크립트 적합성 검토 결과:")

    payload = _extract_first_json(eval_raw) or eval_raw.strip()

    parsed = None
    try:
        parsed = json.loads(payload)
    except Exception:
        parsed = None

    if isinstance(parsed, dict) and parsed.get("compliance") is True:
        print("생성된 스크립트가 시나리오에 부합합니다.")

        # ▼ 여기서 원하는 인자를 지정하세요
        sample_cli_args = ["127.0.0.1:38080", "1234", "sqa-vpes"]  # 질문 주신 3개 인자
        # (선택) 환경변수로도 병행 전달 — default_setting이 env를 읽는 경우 안전망
        sample_env = {
            "VPES_BASE_URL": "http://127.0.0.1:38080",
            "VPES_PASSWORD": "1234",
            "VPES_PROJECT": "sqa-vpes",
        }

        # 자동 실행 + 엑셀 누적
        _ = run_and_log(file_name, cli_args=sample_cli_args, env_vars=sample_env)
    elif isinstance(parsed, dict):
        issues = parsed.get("issues") or []
        if issues:
            print("다음 사유로 시나리오와 불일치합니다:")
            for i, issue in enumerate(issues, 1):
                print(f" - {i}. {issue}")
        else:
            print("시나리오와 완전 일치하지 않는 것으로 보이나, 상세 사유가 제공되지 않았습니다.")
    else:
        print("검토 결과를 해석할 수 없어 원문을 출력합니다:")
        print(eval_raw.strip())


if __name__ == "__main__":
    main()
